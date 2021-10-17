import pandas as pd
import numpy as ny
import csv
import sys, codecs
import copy
import openpyxl

path=sys.argv[1]
with open(path,encoding='utf-8') as f:
    f=csv.reader(f)
    all_csv=[row for row in f]
    csv_r=len(all_csv)
    csv_c=len(all_csv[0])
    print(csv_c,csv_r)
#スパン数
    #xspn=[i for i in all_csv if 'Xスパン数' in i]
    #xs=int(xspn[0][2])
    #yspn=[i for i in all_csv if 'Yスパン数' in i]
    #ys=int(yspn[0][2])
    #ay=xs+1
    #ax=ys+1
    #print(xs,ys)
    #print(ax,ay)
#軸力表摘出
    target="name=支持力検討用軸力表"
    hit= [i for i in all_csv if target in i]
    row_s=all_csv.index(hit[0])
    print(row_s)
    #print(hit)
    row_e=0
    tar="name="
    #result=False
    for i in all_csv[row_s+1:]:
        row_e+=1
        #print(n)
        if tar in i[0]:
            #print(i)
            break
        else:
            continue
    print(row_e)
    #行番号name=支持力検討用軸力表（グリッド形式）
    v_force_list=all_csv[row_s:row_s+row_e-2]#軸力表
    print(v_force_list)

    Xa=[]
    Ya=[]
    f_num=[]
    for i in v_force_list[4:]:
        Xa.append(i[1])
        Ya.append(i[2])
        f_num.append(i[3])
    print(Xa,Ya)

    force_type=v_force_list[2]
    print(force_type)

    Lindex=force_type.index('L')
    Exp_id=force_type.index('L+Ex')
    Exm_id=force_type.index('L-Ex')
    Eyp_id=force_type.index('L+Ey')
    Eym_id=force_type.index('L-Ey')

    L_N=[]
    for i in v_force_list[4:]:
        L_N.append(float(i[Lindex]))
    LpEx=[]
    for i in v_force_list[4:]:
        LpEx.append(float(i[Exp_id]))
    LmEx=[]
    for i in v_force_list[4:]:
        LmEx.append(float(i[Exm_id]))
    LpEy=[]
    for i in v_force_list[4:]:
        LpEy.append(float(i[Eyp_id]))
    LmEy=[]
    for i in v_force_list[4:]:
        LmEy.append(float(i[Eym_id]))

    book=openpyxl.load_workbook("柱状改良の検討.xlsx")
    pex=book['L+Ex']
    mex=book['L-Ex']
    pey=book['L+Ey']
    mey=book['L-Ey']

    print(book.sheetnames)
    def write_list(sheet, in_data, start_row, start_col):
        for x, cell in enumerate(in_data):
            sheet.cell(row=start_row,
                column=start_col + x,
                value=in_data[x])

    write_list(pex, f_num, 3, 5)
    write_list(pex, Xa, 4, 5)
    write_list(pex, Ya, 5, 5)
    write_list(pex, LpEx, 6, 5)

    write_list(mex, f_num, 3, 5)
    write_list(mex, Xa, 4, 5)
    write_list(mex, Ya, 5, 5)
    write_list(mex, LmEx, 6, 5)

    write_list(pey, f_num, 3, 5)
    write_list(pey, Xa, 4, 5)
    write_list(pey, Ya, 5, 5)
    write_list(pey, LpEy, 6, 5)

    write_list(mey, f_num, 3, 5)
    write_list(mey, Xa, 4, 5)
    write_list(mey, Ya, 5, 5)
    write_list(mey, LmEy, 6, 5)

    book.save("柱状改良の検討.xlsx")#sheet=book.worksheets[0]
