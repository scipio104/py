import numpy as np
from matplotlib import pyplot as plt
import csv
import sys, codecs
import copy
import pprint
import pandas as pd
# パラメータを設定
wb = pd.read_excel('固有値.xlsx', 'input')
mx=wb['M']
kx=wb['KX']
ky=wb['KY']
n=list(wb.index.values)
f=len(wb.index.values)
kx=kx.values.tolist()
ky=ky.values.tolist()
floor_num= [(str(i)+"F") for i in range(1,f+1)]
num_f= [(str(i)+"F") for i in range(f,0,-1)]
dim=[(str(i)+"th") for i in range(1,f+1)]
print(wb.index.values)
print(wb.columns.values)
print(wb)
print(mx)
print(kx)
print(ky)
print(floor_num,dim)
print(num_f)
print(type(kx))
m_rev=[]
for i in mx:
    j=round(i*1000,0)
    m_rev.append(j)
m_s=m_rev[::-1]
m_len=len(m_s)
MM =np.square(m_s)
print(m_s)
print(type(m_s))
#Ｍマトリクス作成
m=[]
for i in n:
    m1=[]
    for j in n:
        if i==j:
            m1.append(m_s[i])
        else:
            m1.append(0)
    m.append(m1)
print("----M--(no inv)---")
pd_m=pd.DataFrame(m,columns=dim,index=floor_num)
print(pd_m)
M_inv = np.linalg.inv(m)
M = np.array(M_inv)
#==============================
def k_matrix(input_k):
    k_sort=[]
    k_sort=input_k[::-1]
    k_s_copy=copy.copy(k_sort)
    k_s_copy.append(0)
    del k_s_copy[0]
    k_diagonal=np.array(k_sort) + np.array(k_s_copy)
    matrix=[]
    for i in n:
        k1=[]
        for j in n:
            if i==j:
                k1.append(round(k_diagonal[i],1))
            elif i+1==j:
                k2=round((k_sort[i+1]*-1),1)
                k1.append(k2)
            elif i-1==j:
                k3=round((k_sort[i]*-1),1)
                k1.append(k3)
            else:
                k1.append(0)
        matrix.append(k1)
    return matrix
#==================
def k_unit(matrix_k):
    k_u=[]
    for i in matrix_k:
        ku1=[]
        for j in i:
            j=round(j*1000*1000,0)
            ku1.append(j)
        k_u.append(ku1)
    return k_u
#==================
kx_mat= k_matrix(kx)
ky_mat= k_matrix(ky)
pd_kx=pd.DataFrame(kx_mat,columns=dim,index=floor_num)
pd_ky=pd.DataFrame(ky_mat,columns=dim,index=floor_num)
print("----KX----")
print(pd_kx)
print("----KY----")
print(pd_ky)

KX= np.array(k_unit(kx_mat))
KY= np.array(k_unit(ky_mat))
pd_M=pd.DataFrame(M,columns=dim,index=floor_num)
print(pd_M)

# 固有値と固有ベクトルを計算
x_omega2, vx = np.linalg.eig(np.dot(M, KX))
y_omega2, vy = np.linalg.eig(np.dot(M, KY))
# 固有値の順番を昇順にソート
x_omega_sort= np.sort(x_omega2)
y_omega_sort= np.sort(y_omega2)
x_omega=np.sqrt(x_omega_sort)
y_omega=np.sqrt(y_omega_sort)
# 固有値のソート時のインデックスを取得
# ⇒固有ベクトルと対応させるため
x_sort_index = np.argsort(x_omega2)
y_sort_index = np.argsort(y_omega2)
# 固有値に対応する固有ベクトルをソート
def v_sort(v, sort_index):
    v_rank=[]
    for i in v:
        vr= []
        for j in sort_index:
            vr1=round(i[j],4)
            vr.append(vr1)
        v_rank.append(vr)
    return v_rank
x_v=v_sort(vx,x_sort_index)
y_v=v_sort(vy,y_sort_index)
x_v.reverse()
y_v.reverse()

xv_t_sort = np.array(x_v).T
yv_t_sort = np.array(y_v).T

def period(omega):
    period=[]
    for i in omega:
        period.append(round(2*np.pi/i,3))
    return period
X_period=period(x_omega)
Y_period=period(y_omega)

pd_x_omega2=pd.DataFrame(x_omega_sort,index=dim,columns=["X"]).T
pd_y_omega2=pd.DataFrame(y_omega_sort,index=dim,columns=["Y"]).T
# 結果をコンソール表示

print("----frequency----w2----")
pd.options.display.precision = 3
#pd.options.display.float_format = '{:.4f}'.format
print(pd_x_omega2)#.round(2))
print(pd_y_omega2)#.round(2))
print("----frequency----w----")
pd_x_omega=pd.DataFrame(x_omega,index=dim,columns=["X"]).T
print(pd_x_omega)
pd_y_omega=pd.DataFrame(y_omega,index=dim,columns=["Y"]).T
print(pd_y_omega)
print("----period----")
pd_x_period=pd.DataFrame(X_period,index=dim,columns=["X"]).T
print(pd_x_period)
pd_y_period=pd.DataFrame(Y_period,index=dim,columns=["Y"]).T
print(pd_y_period)
print("----vector----")
pd_X_V=pd.DataFrame(x_v,index=num_f,columns=dim)
pd_Y_V=pd.DataFrame(y_v,index=num_f,columns=dim)
print("----X----")
print(pd_X_V)
print("----Y----")
print(pd_Y_V)

def beata(v_t_sort,m):
    bea=[]
    for i in v_t_sort:
        chi=0
        mot=0
        for j in range(m_len):
            ch=i[j]*m[j]
            mo=i[j]*i[j]*m[j]
            chi=chi+ch
            mot=mot+mo
        bea.append(round(chi/mot,4))
    return bea
x_beata=beata(xv_t_sort,m_s)
y_beata=beata(yv_t_sort,m_s)
print("----sigeki----")
pd_x_beata=pd.DataFrame(x_beata,index=dim,columns=["X"]).T
print(pd_x_beata)
pd_y_beata=pd.DataFrame(y_beata,index=dim,columns=["Y"]).T
print(pd_y_beata)

def bu(beata,v_t_sort):
    bu=[]
    k=-1
    for i in v_t_sort:
        k+=1
        bu1=[]
        for j in range(m_len):
            bu1.append(round(i[j]*beata[j],4))
        bu.append(bu1)
    bu = np.array(bu)
    #bu_sort=np.array(bu).T
    return bu
XBU=bu(x_beata,x_v)
YBU=bu(y_beata,y_v)
print("----shigeki*vector----")
G_xbu=np.array(XBU)
G_ybu=np.array(YBU)
pd.options.display.precision = 6
pd_xbu=pd.DataFrame(XBU,index=num_f,columns=dim)
pd_ybu=pd.DataFrame(YBU,index=num_f,columns=dim)
print(pd_xbu)
print(pd_ybu)

def sum_bu(bu):
    sbu=[]
    for i in range(f):
        k=0
        for j in range(f):
            k=k+bu[i][j]
        sbu.append(round(k,3))
    return sbu
x_sum_bu=sum_bu(XBU)
y_sum_bu=sum_bu(YBU)
print("----sum(shigeki*vector)----")
print(x_sum_bu)
print(y_sum_bu)

#excel
import openpyxl
book=openpyxl.load_workbook("固有値.xlsx")
x=book['X']
y=book['Y']
def write_list(sheet, in_data, start_row, start_col):
    for x, cell in enumerate(in_data):
        sheet.cell(row=start_row,
            column=start_col + x,
            value=in_data[x])
def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])

write_list(x, x_omega_sort, 4, 5)
write_list(x, x_omega, 5, 5)
write_list(x, X_period, 6, 5)
write_list(x, x_beata, 8, 5)
write_list_2d(x, x_v, 10, 5)
write_list_2d(x, XBU, 10+1+f, 5)

write_list(y, y_omega_sort, 4, 5)
write_list(y, y_omega, 5, 5)
write_list(y, Y_period, 6, 5)
write_list(y, y_beata, 8, 5)
write_list_2d(y, y_v, 10, 5)
write_list_2d(y, YBU, 10+1+f, 5)

book.save("固有値.xlsx")
# グラフ化のために自由度軸作成
dof = np.linspace(0, f, f+1)

# ここからグラフ描画
# フォントの種類とサイズを設定する。
plt.rcParams['font.size'] = 14
plt.rcParams['font.family'] = 'Times New Roman'

# 目盛を内側にする。
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

# グラフの上下左右に目盛線を付ける。
fig = plt.figure(figsize=(8,6))
ax1=fig.add_subplot(1,2,1)
ax2=fig.add_subplot(1,2,2)

ax1.yaxis.set_ticks_position('both')
ax1.xaxis.set_ticks_position('both')
ax2.yaxis.set_ticks_position('both')
ax2.xaxis.set_ticks_position('both')

# 軸のラベルを設定する。
ax1.set_ylabel('Degree of freedom')
ax1.set_xlabel('X_Eigenvector')
ax2.set_xlabel('Y_Eigenvector')
# データの範囲と刻み目盛を明示する。
ax1.set_xticks(np.arange(-2, 2, 0.5))
ax1.set_yticks(np.arange(0, f+2, 1))
ax1.set_xlim(-1.5,1.5)
ax1.set_ylim(0, f+2)

ax2.set_xticks(np.arange(-2, 2, 0.5))
ax2.set_yticks(np.arange(0, f+2, 1))
ax2.set_xlim(-1.5,1.5)
ax2.set_ylim(0, f+2)
# データプロット 固有ベクトルの形を見る
for i in range(f):
    eigen_vector = np.concatenate([[0],XBU[i]])
    ax1.plot(eigen_vector, dof, lw=1, marker='o')
for i in range(f):
    eigen_vector = np.concatenate([[0],YBU[i]])
    ax2.plot(eigen_vector, dof, lw=1, marker='o')

fig.tight_layout()
plt.legend()
# グラフを表示する。

plt.show()

plt.close()
