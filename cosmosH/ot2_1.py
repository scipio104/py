import pandas as pd
import numpy as ny
import csv
import sys, codecs
import copy
import openpyxl

path=sys.argv[1]
with open(path,encoding='utf-8') as f:
    f=csv.reader(f)
    all_csv=[row for row in f]
    csv_r=len(all_csv)
    csv_c=len(all_csv[0])
    print(csv_c,csv_r)
#軸力表摘出
    #start_row
    target1="name=基礎設計用軸力表"
    hit1= [i for i in all_csv if target1 in i]
    row_fs=all_csv.index(hit1[0])
    #end_row
    tar="name="
    row_fe=0
    for i in all_csv[row_fs+1:]:
        row_fe+=1
        if tar in i[0]:
            break
        else:
            continue
    print(row_fs,row_fe)
#行番号name=基礎設計用軸力表
    #f_weight_list=all_csv[row_fs:row_fs+row_fe-2]#軸力表
    #print(f_weight_list)

    #行番号name=支持力検討用軸力表（グリッド形式）
    v_force_list=all_csv[row_fs:row_fs+row_fe-2]#軸力表
    print(v_force_list)

    Xa=[]
    Ya=[]
    f_num=[]
    for i in v_force_list[4:]:
        Xa.append(i[1])
        Ya.append(i[2])
        f_num.append(i[3])
    print(Xa,Ya,f_num)



    print('====')
    force_type=v_force_list[2]
    print(force_type)

    Lindex=force_type.index('L')
    Exp_id=force_type.index('L+Ex')
    Exm_id=force_type.index('L-Ex')
    Eyp_id=force_type.index('L+Ey')
    Eym_id=force_type.index('L-Ey')

    L_N=[]
    for i in v_force_list[4:]:
        L_N.append(float(i[Lindex]))
    LpEx=[]
    for i in v_force_list[4:]:
        LpEx.append(float(i[Exp_id]))
    LmEx=[]
    for i in v_force_list[4:]:
        LmEx.append(float(i[Exm_id]))
    LpEy=[]
    for i in v_force_list[4:]:
        LpEy.append(float(i[Eyp_id]))
    LmEy=[]
    for i in v_force_list[4:]:
        LmEy.append(float(i[Eym_id]))

#MAX軸力
    f_name=list(set(f_num))
    f_nn=len(f_name)
    col_f=len(f_num)
    print(f_name,f_nn,col_f)

    m1=[]
    for i in range(f_nn):
        m2=[]
        for j in range(col_f):
            if f_name[i]==f_num[j]:
                m2.append(L_N[j])
        m1.append(m2)

    fmax=[]
    for i in m1:
        fmax.append(max(i))

    m4=[]
    for i in m1:
        m4.append(max(i))
    m5=[]
    for i in m4:
        m5.append(L_N.index(i))
    print(m5)

    Xmax=[]
    Ymax=[]
    f_num=[]
    for i in m5:
        Xmax.append(Xa[i])
        Ymax.append(Ya[i])
        #f_num.append(i[3])
    print(Xmax,Ymax,fmax)
    position=[]
    for i,j in zip(Xmax,Ymax):
        position.append(i+'-'+j)
    print(position)


    book=openpyxl.load_workbook("柱状改良の検討.xlsx")
    pex=book['L+Ex']
    mex=book['L-Ex']
    pey=book['L+Ey']
    mey=book['L-Ey']



    print(book.sheetnames)
    def write_list(sheet, in_data, start_row, start_col):
        for x, cell in enumerate(in_data):
            sheet.cell(row=start_row,
                column=start_col + x,
                value=in_data[x])

    write_list(pex, f_num, 3, 5)
    write_list(pex, Xa, 4, 5)
    write_list(pex, Ya, 5, 5)
    write_list(pex, LpEx, 6, 5)

    write_list(mex, f_num, 3, 5)
    write_list(mex, Xa, 4, 5)
    write_list(mex, Ya, 5, 5)
    write_list(mex, LmEx, 6, 5)

    write_list(pey, f_num, 3, 5)
    write_list(pey, Xa, 4, 5)
    write_list(pey, Ya, 5, 5)
    write_list(pey, LpEy, 6, 5)

    write_list(mey, f_num, 3, 5)
    write_list(mey, Xa, 4, 5)
    write_list(mey, Ya, 5, 5)
    write_list(mey, LmEy, 6, 5)

    write_list('鉛直', position, 7, 15)




    book.save("柱状改良の検討.xlsx")#sheet=book.worksheets[0]
