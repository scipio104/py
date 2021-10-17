import pandas as pd
import numpy as ny
import csv
import sys, codecs
import copy
import openpyxl

path=sys.argv[1]
with open(path,encoding='utf-8') as f:
    f=csv.reader(f)
    all_csv=[row for row in f]
    csv_r=len(all_csv)
    csv_c=len(all_csv[0])
    print(csv_c,csv_r)
#基礎自重
    target1="name=基礎自重"
    hit1= [i for i in all_csv if target1 in i]
    row_fs=all_csv.index(hit1[0])
    print(row_fs)
    #print(hit)
    row_fe=0
    tar="name="
    for i in all_csv[row_fs+1:]:
        row_fe+=1
        #print(n)
        if tar in i[0]:
            print(i)
            break
        else:
            continue
    print(row_fe)
    #行番号name=支持力検討用軸力表（グリッド形式）
    fund_list=all_csv[row_fs:row_fs+row_fe-2]#軸力表
    print(fund_list)#スパン数

    Af=[]
    Df=[]
    Wf=[]
    for i in fund_list[4:]:
        Af.append(float(i[4]))
        Df.append(float(i[5]))
        Wf.append(float(i[6]))
    print(Af,Df,Wf)    #xspn=[i for i in all_csv if 'Xスパン数' in i]
    #xs=int(xspn[0][2])
    #yspn=[i for i in all_csv if 'Yスパン数' in i]
    #ys=int(yspn[0][2])
    #ay=xs+1
    #ax=ys+1
    #print(xs,ys)
    #print(ax,ay)
#軸力表摘出
    target="name=支持力検討用軸力表"
    hit= [i for i in all_csv if target in i]
    row_s=all_csv.index(hit[0])
    print(row_s)
    #print(hit)
    row_e=0
    tar="name="
    #result=False
    for i in all_csv[row_s+1:]:
        row_e+=1
        #print(n)
        if tar in i[0]:
            #print(i)
            break
        else:
            continue
    print(row_e)
    #行番号name=支持力検討用軸力表（グリッド形式）
    v_force_list=all_csv[row_s:row_s+row_e-2]#軸力表
    print(v_force_list)

    Xa=[]
    Ya=[]
    f_num=[]
    for i in v_force_list[4:]:
        Xa.append(i[1])
        Ya.append(i[2])
        f_num.append(i[3])
    print(Xa,Ya)
    f_name=list(set(f_num))


    print('======')
    print(type(f_name))

    print(f_name)

    force_type=v_force_list[2]
    print(force_type)

    Lindex=force_type.index('L')
    Exp_id=force_type.index('L+Ex')
    Exm_id=force_type.index('L-Ex')
    Eyp_id=force_type.index('L+Ey')
    Eym_id=force_type.index('L-Ey')

    L_N=[]
    for i in v_force_list[4:]:
        L_N.append(float(i[Lindex]))
    LpEx=[]
    for i in v_force_list[4:]:
        LpEx.append(float(i[Exp_id]))
    LmEx=[]
    for i in v_force_list[4:]:
        LmEx.append(float(i[Exm_id]))
    LpEy=[]
    for i in v_force_list[4:]:
        LpEy.append(float(i[Eyp_id]))
    LmEy=[]
    for i in v_force_list[4:]:
        LmEy.append(float(i[Eym_id]))

    l_zip = [(s1, s2) for s1, s2 in zip(f_num,L_N)]
    print(l_zip)
    fn=len(f_name)
    lz=len(l_zip)

    print(fn,lz)



    book=openpyxl.load_workbook("柱状改良の検討.xlsx")
    long=book['L']
    pex=book['L+Ex']
    mex=book['L-Ex']
    pey=book['L+Ey']
    mey=book['L-Ey']

    print(book.sheetnames)
    def write_list(sheet, in_data, start_row, start_col):
        for x, cell in enumerate(in_data):
            sheet.cell(row=start_row,
                column=start_col + x,
                value=in_data[x])

    write_list(long, f_num, 8, 5)
    write_list(long, Xa, 9, 5)
    write_list(long, Ya, 10, 5)
    write_list(long, L_N, 11, 5)
    write_list(long, Af, 24, 5)

    write_list(pex, f_num, 8, 5)
    write_list(pex, Xa, 9, 5)
    write_list(pex, Ya, 10, 5)
    write_list(pex, LpEx, 11, 5)
    write_list(pex, Af, 18, 5)

    write_list(mex, f_num, 8, 5)
    write_list(mex, Xa, 9, 5)
    write_list(mex, Ya, 10, 5)
    write_list(mex, LmEx, 11, 5)
    write_list(mex, Af, 18, 5)

    write_list(pey, f_num, 8, 5)
    write_list(pey, Xa, 9, 5)
    write_list(pey, Ya, 10, 5)
    write_list(pey, LpEy, 11, 5)
    write_list(pey, Af, 18, 5)

    write_list(mey, f_num, 8, 5)
    write_list(mey, Xa, 9, 5)
    write_list(mey, Ya, 10, 5)
    write_list(mey, LmEy, 11, 5)
    write_list(mey, Af, 18, 5)

    book.save("柱状改良の検討.xlsx")#sheet=book.worksheets[0]
